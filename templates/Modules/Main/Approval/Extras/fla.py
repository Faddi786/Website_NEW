from flask import Flask
import openpyxl

app = Flask(__name__)

@app.route('/')
def extract_manager_data():
    # Load the Excel file
    wb = openpyxl.load_workbook('managers.xlsx')
    sheet = wb.active
    
    # Get values from the first row
    first_row = [cell.value for cell in sheet[1]]
    
    # Check if 'Manager1' exists in the first row
    if 'Manager1' in first_row:
        # Find the index of 'Manager1' column
        manager1_index = first_row.index('Manager1') + 1  # Adding 1 because index starts from 0
        
        # Extract all values from the 'Manager1' column
        manager1_data = [sheet.cell(row=row, column=manager1_index).value for row in range(2, sheet.max_row + 1)]
        
        # Call find_employee_data function and return its result
        print(find_employee_data(manager1_data))
        return "find_employee_data(manager1_data)"
    else:
        return "Manager1 column not found"


def find_employee_data(sheet_names):
    # Convert the list of sheet names to a comma-separated string
    sheet_names_str = ','.join(sheet_names)
    
    # Load the Excel file
    wb = openpyxl.load_workbook('employee_data.xlsx')
    
    # Initialize a list to store data from matching sheets
    matching_sheets_data = []
    
    # Iterate through each sheet in the Excel file
    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            # If the sheet exists, get its data
            sheet = wb[sheet_name]
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                sheet_data.append(row)
            matching_sheets_data.append({sheet_name: sheet_data})
    
    # If no matching sheets found, return a message
    if not matching_sheets_data:
        return "No matching sheets found"
    else:
        return str(matching_sheets_data)


if __name__ == '__main__':
    app.run(debug=True)
