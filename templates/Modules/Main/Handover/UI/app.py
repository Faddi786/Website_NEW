from flask import Flask, jsonify
from openpyxl import load_workbook
from flask import Flask, request, render_template, jsonify
import pandas as pd
import re
from datetime import date
from dateutil import tz
from datetime import datetime as dt, timezone
import datetime
from flask import Flask, request, render_template, jsonify
import subprocess
import json
import pandas as pd
from flask_cors import CORS  # Import CORS
from flask_cors import cross_origin
from itertools import count
import re
import os
import openpyxl
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from flask import Flask, request, redirect, url_for
from bs4 import BeautifulSoup
from flask import session
from datetime import date
from dateutil import tz
from datetime import datetime as dt, timezone
app = Flask(__name__)

@app.route('/')
def index():
    return render_template('shopping.html')


@app.route('/cart_items')
def cart_items():
    # Load the Excel file
    wb = load_workbook('Excel/inventory.xlsx')
    sheet = wb.active

    # Define a list to store the dictionaries
    data = []

    # Iterate over rows and append data to the list as dictionaries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {
            'Product Name': row[0],
            'Make': row[2],
            'Model': row[3],
            'Product Serial': row[4],
            'productCategory': row[1],
            'Current Owner': row[8],
            'Current Project': row[9],
        }

        data.append(item)
    print(data)
    return jsonify(data=data)


@app.route('/submit_form', methods=['POST'])
def submit_form():
    form_data = request.json  # Get JSON data from the request
    print("This is the form data",form_data)
    # Mapping of keys to column indices
    key_to_column = {
        'Category': 1,
        'Name': 2,
        'Make': 3,
        'Model': 4,
        'Product_Serial': 5,
        'Condition': 11,
        'SenderFrontend': 13,
        'SenderBackend': 14,
        'SenderCleanup': 15,
        'SenderRemarks': 16,
    }

    # Define column names
    columns = ['Serial No', 'Category', 'Name', 'Make', 'Model', 'Product_Serial', 'Handover location', 'Handover authorized by', 'Handover date', 'Receiver Date', 'Acceptance Date', 'Condition', 'Reached', 'SenderFrontend', 'SenderBackend', 'SenderCleanup', 'SenderRemarks', 'ReceiverFrontend', 'ReceiverBackend', 'ReceiverCleanup', 'ReceiverRemarks']

    # Get the name from session data
    name = "Fahad"

    # Load existing Excel file if it exists, otherwise create an empty DataFrame
    try:
        existing_df = pd.read_excel('Excel/handover_data.xlsx', sheet_name=str(name))
    except FileNotFoundError:
        existing_df = pd.DataFrame(columns=columns)
    except ValueError:
        existing_df = pd.DataFrame(columns=columns)

    # Create a DataFrame from the form data
    new_data = []
    for item in form_data[1:]:  # Exclude the first row which contains metadata
        new_row = [None] * len(columns)  # Initialize a new row with None values
        for key, value in item.items():
            if key in key_to_column:
                if isinstance(key_to_column[key], int):
                    new_row[key_to_column[key]] = value
                elif isinstance(key_to_column[key], dict):
                    # Set flags for 'frontend', 'backend', and 'cleanup' based on the 'Status' field
                    status_parts = value.split(',')
                    for status_part in status_parts:
                        if status_part.strip() in key_to_column[key]:
                            new_row[key_to_column[key][status_part.strip()]] = 1
        new_data.append(new_row)

    new_df = pd.DataFrame(new_data, columns=columns)

    # Determine the next value for "Serial No" for the new data
    if not existing_df.empty:
        last_serial_no = existing_df['Serial No'].iloc[-1]  # Get the last serial number
        last_serial_no = str(last_serial_no)  # Convert to string explicitly
        match = re.match(r"(\d+)\.(\d+)", last_serial_no)  # Extract numerical parts using regex
        last_main_num = int(match.group(1))  # Extract the main number part
        last_sub_num = int(match.group(2))  # Extract the sub number part
    else:
        last_main_num = 0
        last_sub_num = 0

    # Calculate the total number of rows in the existing DataFrame and new DataFrame combined
    total_rows = len(existing_df) + len(new_df)

    # Determine the current numerical part for the new batch
    current_main_num = last_main_num + 1

    # Reset the subpart if the main part changes
    if current_main_num > last_main_num:
        last_sub_num = 0

    # Iterate over each row of new data and generate serial numbers
    serial_numbers = []
    for i in range(len(new_df)):
        current_sub_num = last_sub_num + i + 1
        new_serial_no = str(current_main_num) + '.' + str(current_sub_num)
        serial_numbers.append(new_serial_no)

    # Add serial numbers to the new data frame
    new_df['Serial No'] = serial_numbers

    # Add current date to the new data frame
    new_df['Acceptance Date'] = datetime.now().date()

    # Concatenate existing and new data frames
    df = pd.concat([existing_df, new_df], ignore_index=True)

    # Write the updated data frame back to the Excel file
    with pd.ExcelWriter('Excel/handover_data.xlsx', engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name=str(name))

    # Return a success message
    return jsonify({'message': 'Excel file updated successfully'})




if __name__ == '__main__':
    app.run(debug=True, port=5001)
