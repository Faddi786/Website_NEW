from flask import Flask, jsonify, session
from openpyxl import load_workbook
from flask import Flask, request, render_template, jsonify
import pandas as pd
import random 
from flask import jsonify
from openpyxl import load_workbook
import pandas as pd
import re 

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('handover.html')



def extract_data_from_excel(product_ids):
    # Load the Excel file
    df = pd.read_excel('Excel/handover_data.xlsx')
    
    # Initialize a list to store the extracted data
    extracted_data = []
    
    # Iterate over each product id
    for product_id in product_ids:
        # Filter rows where 'ProductID' column matches the product id
        filtered_rows = df[df['ProductID'] == product_id]
        
        # Iterate over each filtered row
        for index, row in filtered_rows.iterrows():
            # Extract the values from the row and append to the extracted_data list
            extracted_data.append([row['ProductID'], row['Name'], row['FormID'], row['HandoverDate']])
    
    return extracted_data

@app.route('/cart_items')
def cart_items():
    # Load the Excel file
    wb = load_workbook('Excel/inventory.xlsx')
    sheet = wb.active

    # Define a list to store the dictionaries
    data = []

    # Define a list to store ProductIDs
    product_ids = []

    # Get the name from session data
    name = "Fahad"
    
    # Iterate over rows and append data to the list as dictionaries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if the current owner matches the name from the session cookie
        if row[8] == name:
            # Append the ProductID to the list
            product_ids.append(row[5])
            
            item = {
                'Category': row[1],
                'ProductID': row[5],
                'Name': row[2],
                'Make': row[3],
                'Model': row[4],
                'FromPerson' : row[8]
            }
            data.append(item)

    # Call extract_data_from_excel function to get result
    result = extract_data_from_excel(product_ids)
    
    # Replace NaN values with "NAN" in data
    data = [{k: "NAN" if pd.isna(v) else v for k, v in item.items()} for item in data]

    # Replace NaN values with "NAN" in result
    result = [[col if pd.notna(col) else "NAN" for col in row] for row in result]

    # Create a new list containing both data and result
    combined_data = [data, result]
    #print("this is the combined data",combined_data)
    return jsonify(combined_data=combined_data)


global from_person 
global to_person 
global from_project 
global to_project 


@app.route('/send_approval_request', methods=['POST'])
def send_approval_request():
    
    global from_person 
    global to_person 
    global from_project 
    global to_project 

    form_data = request.json  # Get JSON data from the request
    print("This is the form data for approval request", form_data)

    # Check if form_data is not empty and is a list
    if form_data and isinstance(form_data, list) and len(form_data) > 0:
        first_dict = form_data[0]  # Get the first dictionary from the list

        # Access and print values of 'FromPerson', 'ToPerson', 'FromProject', and 'ToProject'
        from_person = first_dict.get('FromPerson')
        to_person = first_dict.get('ToPerson')
        from_project = first_dict.get('FromProject')
        to_project = first_dict.get('ToProject')

        print("From Person:", from_person)
        print("To Person:", to_person)
        print("From Project:", from_project)
        print("To Project:", to_project)
    else:
        print("No form data or invalid data format received")


    # Mapping of keys to column indices
    key_to_column = {
        'FormID': 0,  # Add FormID to the mapping
        'Category': 2,
        'ProductID': 3,
        'Name': 4,
        'Make': 5,
        'Model': 6,
        'SenderCondition': 11,
        'SenderRemarks': 12,
    }

    # Define column names
    columns = ['FormID', 'Serial', 'Category', 'ProductID', 'Name', 'Make', 'Model', 
                'FromProject', 'ToProject','FromPerson', 'ToPerson',  
                'SenderCondition', 'SenderRemarks', 'ReceiverCondition', 'ReceiverRemarks' ]  # Add FormID to the columns list

    # Load existing Excel file if it exists, otherwise create an empty DataFrame
    try:
        existing_df = pd.read_excel('Excel/handover_data.xlsx')
    except FileNotFoundError:
        existing_df = pd.DataFrame(columns=columns)
    except ValueError:
        existing_df = pd.DataFrame(columns=columns)

    # Create a DataFrame from the form data
    new_data = []
    for item in form_data[1:]:  # Exclude the first row which contains metadata
        new_row = [None] * len(columns)  # Initialize a new row with None values
        for key, value in item.items():
            if key in key_to_column:
                if isinstance(key_to_column[key], int):
                    new_row[key_to_column[key]] = value
                elif isinstance(key_to_column[key], dict):
                    # Set flags for 'frontend', 'backend', and 'cleanup' based on the 'Status' field
                    status_parts = value.split(',')
                    for status_part in status_parts:
                        if status_part.strip() in key_to_column[key]:
                            new_row[key_to_column[key][status_part.strip()]] = 1
        new_data.append(new_row)

    new_df = pd.DataFrame(new_data, columns=columns)

    # Generate unique FormID
    def generate_form_id():
        original_id = 'abcd1234'
        id_list = list(original_id)
        random.shuffle(id_list)
        return ''.join(id_list)

    # Extract values from the "FormID" column
    form_ids = existing_df['FormID'].tolist()

    # Generate a unique FormID
    unique_form_id = generate_form_id()
    while unique_form_id in form_ids:
        unique_form_id = generate_form_id()

    # Insert the unique FormID into the "FormID" column of the first product
    if len(new_df) > 0:
        new_df.loc[0, 'FormID'] = unique_form_id

    if len(new_df) > 0:
        new_df.loc[0, 'FromProject'] = from_project
    
    if len(new_df) > 0:
        new_df.loc[0, 'ToProject'] = to_project
    
    if len(new_df) > 0:
        new_df.loc[0, 'FromPerson'] = from_person

    if len(new_df) > 0:
        new_df.loc[0, 'ToPerson'] = to_person

    text="Send Form"
    # Call the function to send email with the generated PDF attached
    send_email(text, unique_form_id ,None,from_person,to_person,from_project ,to_project )


    # Determine the next value for "Serial No" for the new data
    if not existing_df.empty and 'Serial' in existing_df:
        last_serial_no = existing_df['Serial'].iloc[-1]  # Get the last serial number
        last_serial_no = str(last_serial_no)  # Convert to string explicitly
        match = re.match(r"(\d+)\.(\d+)", last_serial_no)  # Extract numerical parts using regex
        if match:
            last_main_num = int(match.group(1))  # Extract the main number part
            last_sub_num = int(match.group(2))  # Extract the sub number part
        else:
            last_main_num = 0
            last_sub_num = 0
    else:
        last_main_num = 0
        last_sub_num = 0

    # Calculate the total number of rows in the existing DataFrame and new DataFrame combined
    total_rows = len(existing_df) + len(new_df)

    # Determine the current numerical part for the new batch
    current_main_num = last_main_num + 1

    # Reset the subpart if the main part changes or if there are no existing serial numbers
    if current_main_num > last_main_num or total_rows == 0:
        last_sub_num = 0

    # Iterate over each row of new data and generate serial numbers
    serial_numbers = []
    for i in range(len(new_df)):
        current_sub_num = last_sub_num + i + 1
        new_serial_no = str(current_main_num) + '.' + str(current_sub_num)
        serial_numbers.append(new_serial_no)

    # Add serial numbers to the new data frame
    new_df['Serial'] = serial_numbers

    # Concatenate existing and new data frames
    df = pd.concat([existing_df, new_df], ignore_index=True)

    # Write the updated data frame back to the Excel file
    with pd.ExcelWriter('Excel/handover_data.xlsx', engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False)

    # Return a success message
    return jsonify({'message': 'Excel file updated successfully'})




def send_email(message_content, form_no, eway_bill_no=None,from_person=None,to_person=None,from_project =None ,to_project =None ):

    print("This is the email From Person:", from_person)
    print("This is the email To Person:", to_person)
    print("This is the email From Project:", from_project)
    print("This is the email To Project:", to_project)

    # Sender and receiver email addresses
    sender_email = "shaikhfahad687@gmail.com"  # Update with your Gmail address
    receiver_email = "shaikhfahad687@gmail.com"  # Update with the receiver's email address

    try:
        # Gmail SMTP server details
        smtp_server = "smtp.gmail.com"
        smtp_port = 587  # Port for TLS encryption

        # Login credentials (use app password)
        email_password = "rgbbdlhpyleheico"  # Update with your Gmail app password

        # Establish a secure connection with the SMTP server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            # Login to the SMTP server
            server.login(sender_email, email_password)
            
            # Create a multipart message
            message = MIMEMultipart()
            message["From"] = sender_email
            message["To"] = receiver_email
            
            if message_content == "Send Form":
                message["Subject"] = "Handover Transaction details"
                body_message = f"I want to send items, The excel sheet of the eway bill is attached below. Please generate the eway bill and approve this form. The form details are as follows:\n\nForm no: {form_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))
                
                # Get the current working directory
                current_directory = os.getcwd()
                # Assuming the Excel sheet is in the same directory, change the filename if it's different
                excel_file_path = os.path.join(current_directory, "eway_bill.xlsx")

                # Open the Excel file in binary mode
                with open(excel_file_path, "rb") as attachment:
                    # Add Excel file as application/octet-stream
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(attachment.read())

                # Encode file in ASCII characters to send by email    
                encoders.encode_base64(part)

                # Add header as key/value pair to attachment part
                part.add_header("Content-Disposition", f"attachment; filename= {os.path.basename(excel_file_path)}")

                # Add attachment to message
                message.attach(part)
                
            elif message_content == "Receive Form":
                message["Subject"] = "Receive Form details"
                body_message = f"I have received the items, need your approval to proceed with using them, please approve this form. The form details are as follows:\n\nForm no: {form_no}\nEway Bill no: {eway_bill_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"
                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))
                
            elif message_content == "Send Approval Form":
                message["Subject"] = "Send Approval Form details"
                body_message = f"I have granted permission to transfer items. The form details are as follows:\n\nForm no: {form_no}\nEway Bill no: {eway_bill_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"
                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))
                
            elif message_content == "Receive Approval Form":
                message["Subject"] = "Receive Approval Form details"
                body_message = f"I have granted permission to accept the items received. The form details are as follows:\n\nForm no: {form_no}\nEway Bill no: {eway_bill_no}\nFrom Project: {from_project}\nTo Project: {to_project}\nFrom Person: {from_person}\nTo Person: {to_person}"

                
                # Add body to email
                message.attach(MIMEText(body_message, "plain"))

            # Convert the message to string
            str_message = message.as_string()
            
            # Send email
            server.sendmail(sender_email, receiver_email, str_message)
    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    app.run(debug=True, port=5001)
