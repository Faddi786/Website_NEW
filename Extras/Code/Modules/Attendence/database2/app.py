from flask import Flask, request, render_template
import openpyxl
import pandas as pd

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('database.html')

@app.route('/display-data', methods=['POST'])
def display_data():
    data = request.json
    year = data['year']
    month = data['month']
    name = data['name']
    
    file_path = f"{year}.xlsx"  # Path to your Excel file
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        return f"Excel file for year {year} not found."
    
    sheet_names = workbook.sheetnames
    if month not in sheet_names:
        return f"No sheet found for the month {month} in the Excel file."
    
    sheet = workbook[month]
    
    # Initialize an empty DataFrame to store found records
    found_records = pd.DataFrame(columns=['Row', 'Name'])
    
    for row_number, row in enumerate(sheet.iter_rows(), start=1):
        if row[0].value == name:
            found_records = found_records._append({'Row': row_number, 'Name': name}, ignore_index=True)
    
    if found_records.empty:
        print(f"No record found for {name} in {month} of {year}.")
        return f"No record found for {name} in {month} of {year}."
    
    # Print the DataFrame
    print(found_records)
    
    return "Records found, check console for details."

if __name__ == '__main__':
    app.run(debug=True)
