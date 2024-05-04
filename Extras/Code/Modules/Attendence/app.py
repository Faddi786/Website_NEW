from flask import Flask, send_file, render_template
from openpyxl import load_workbook, Workbook
from datetime import datetime

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('emp.html')

def get_sheet_name(date):
    return date.strftime('%B')

def create_or_get_sheet(wb, date):
    sheet_name = get_sheet_name(date)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)
    return wb[sheet_name]

def get_workbook_name(date):
    return f"{date.year}.xlsx"

def get_workbook():
    current_date = datetime.now()
    workbook_name = get_workbook_name(current_date)
    try:
        wb = load_workbook(workbook_name)
    except FileNotFoundError:
        wb = Workbook()
    return wb

@app.route('/get-attendance', methods=['POST'])
def get_attendance():
    # Get the current date
    current_date = datetime.now()
    
    # Get or create workbook for the current year
    wb = get_workbook()
    
    # Get or create sheet for the current month
    sheet = create_or_get_sheet(wb, current_date)
    
    # Find the last used column in the specified row
    day_of_month = current_date.day
    last_column = 1
    while sheet.cell(row=day_of_month, column=last_column).value is not None:
        last_column += 1
    
    # Add 'Tom' to the next available column in the specified row
    sheet.cell(row=day_of_month, column=last_column, value='Tom')
    
    # Save the Excel file
    workbook_name = get_workbook_name(current_date)
    wb.save(workbook_name)
    
    # Send the file to the client
    return 'Attendance recorded successfully'

if __name__ == '__main__':
    app.run(debug=True, port=5001)
